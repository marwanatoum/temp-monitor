"""
Temp Monitor — نظام مراقبة أجهزة صناعية متعددة (Prototype)
-------------------------------------------------------------
- إدارة أجهزة يدوية (إضافة/حذف) مع عتبات حرارة وإحداثيات و tags مخصصة
- لوحة تحكم لجهاز واحد مع منتقي تاريخ/وقت
- خريطة لعرض مواقع الأجهزة
- تصدير CSV/Excel بمدى زمني محدد
"""

import os
import io
import csv
import json
import sqlite3
from datetime import datetime, timezone
from flask import Flask, request, jsonify, render_template, g, redirect, url_for, flash, Response
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

APP_DB = os.path.join(os.path.dirname(__file__), "readings.db")
DEVICE_API_KEY = os.environ.get("DEVICE_API_KEY", "changeme-esp32-key")
DEFAULT_USERNAME = os.environ.get("DEFAULT_USERNAME", "admin")
DEFAULT_PASSWORD = os.environ.get("DEFAULT_PASSWORD", "admin123")

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")

login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message = "الرجاء تسجيل الدخول للمتابعة"


class User(UserMixin):
    def __init__(self, row):
        self.id = str(row["id"])
        self.username = row["username"]


@login_manager.user_loader
def load_user(user_id):
    conn = sqlite3.connect(APP_DB)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    return User(row) if row else None


# ---------------------------------------------------------------------------
# قاعدة البيانات
# ---------------------------------------------------------------------------
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(APP_DB)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    conn = sqlite3.connect(APP_DB)
    conn.execute("PRAGMA foreign_keys = ON")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS devices (
            device_id TEXT PRIMARY KEY,
            description TEXT,
            lat REAL,
            lng REAL,
            temp_min REAL,
            temp_max REAL,
            created_at TEXT NOT NULL
        )
        """
    )

    existing_device_cols = [r[1] for r in conn.execute("PRAGMA table_info(devices)").fetchall()]
    for col, col_type in [
        ("description", "TEXT"),
        ("lat", "REAL"),
        ("lng", "REAL"),
        ("temp_min", "REAL"),
        ("temp_max", "REAL"),
    ]:
        if col not in existing_device_cols:
            conn.execute(f"ALTER TABLE devices ADD COLUMN {col} {col_type}")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS device_tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT NOT NULL REFERENCES devices(device_id) ON DELETE CASCADE,
            tag_name TEXT NOT NULL,
            unit TEXT,
            min_val REAL,
            max_val REAL,
            sort_order INTEGER DEFAULT 0
        )
        """
    )
    existing_tag_cols = [r[1] for r in conn.execute("PRAGMA table_info(device_tags)").fetchall()]
    for col, col_type in [("min_val", "REAL"), ("max_val", "REAL")]:
        if col not in existing_tag_cols:
            conn.execute(f"ALTER TABLE device_tags ADD COLUMN {col} {col_type}")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS readings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            device_id TEXT NOT NULL,
            temperature REAL,
            humidity REAL,
            values_json TEXT,
            alarm INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )

    # ترقية تلقائية: إذا كان جدول readings موجوداً من نسخة قديمة وناقصه أعمدة، نضيفها
    existing_cols = [r[1] for r in conn.execute("PRAGMA table_info(readings)").fetchall()]
    for col, col_type in [
        ("temperature", "REAL"),
        ("humidity", "REAL"),
        ("values_json", "TEXT"),
        ("alarm", "INTEGER DEFAULT 0"),
    ]:
        if col not in existing_cols:
            conn.execute(f"ALTER TABLE readings ADD COLUMN {col} {col_type}")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL
        )
        """
    )

    existing = conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()[0]
    if existing == 0:
        conn.execute(
            "INSERT INTO users (username, password_hash) VALUES (?, ?)",
            (DEFAULT_USERNAME, generate_password_hash(DEFAULT_PASSWORD)),
        )
        print(f"[init] تم إنشاء مستخدم افتراضي: {DEFAULT_USERNAME} / {DEFAULT_PASSWORD}")

    conn.commit()
    conn.close()


def device_to_dict(row, tags):
    d = dict(row)
    d["tags"] = tags
    return d


def get_device_tags(db, device_id):
    rows = db.execute(
        "SELECT tag_name, unit, min_val, max_val FROM device_tags WHERE device_id = ? ORDER BY sort_order ASC, id ASC",
        (device_id,),
    ).fetchall()
    return [
        {"tag_name": r["tag_name"], "unit": r["unit"], "min_val": r["min_val"], "max_val": r["max_val"]}
        for r in rows
    ]


# ---------------------------------------------------------------------------
# تسجيل الدخول / الخروج
# ---------------------------------------------------------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        db = get_db()
        row = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()

        if row and check_password_hash(row["password_hash"], password):
            login_user(User(row))
            return redirect(url_for("dashboard"))

        flash("اسم المستخدم أو كلمة المرور غير صحيحة")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# الصفحات
# ---------------------------------------------------------------------------
@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html", username=current_user.username)


@app.route("/devices")
@login_required
def devices_page():
    return render_template("devices.html", username=current_user.username)


@app.route("/map")
@login_required
def map_page():
    return render_template("map.html", username=current_user.username)


# ---------------------------------------------------------------------------
# API: إدارة الأجهزة (CRUD)
# ---------------------------------------------------------------------------
@app.route("/api/devices", methods=["GET"])
@login_required
def list_devices():
    db = get_db()
    rows = db.execute("SELECT * FROM devices ORDER BY created_at DESC").fetchall()
    result = [device_to_dict(r, get_device_tags(db, r["device_id"])) for r in rows]
    return jsonify(result)


@app.route("/api/devices", methods=["POST"])
@login_required
def create_device():
    data = request.get_json(silent=True) or {}
    device_id = (data.get("device_id") or "").strip()

    if not device_id:
        return jsonify({"error": "device_id مطلوب"}), 400

    db = get_db()
    existing = db.execute("SELECT 1 FROM devices WHERE device_id = ?", (device_id,)).fetchone()
    if existing:
        return jsonify({"error": "هذا الجهاز موجود بالفعل"}), 409

    def _to_float(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    db.execute(
        """
        INSERT INTO devices (device_id, description, lat, lng, created_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (
            device_id,
            data.get("description") or None,
            _to_float(data.get("lat")),
            _to_float(data.get("lng")),
            datetime.now(timezone.utc).isoformat(),
        ),
    )

    tags = data.get("tags") or []
    for i, tag in enumerate(tags):
        name = (tag.get("tag_name") or "").strip()
        if not name:
            continue
        db.execute(
            "INSERT INTO device_tags (device_id, tag_name, unit, min_val, max_val, sort_order) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                device_id, name, (tag.get("unit") or "").strip() or None,
                _to_float(tag.get("min_val")), _to_float(tag.get("max_val")), i,
            ),
        )

    db.commit()
    return jsonify({"status": "ok"}), 201


@app.route("/api/devices/<device_id>", methods=["GET"])
@login_required
def get_device(device_id):
    db = get_db()
    row = db.execute("SELECT * FROM devices WHERE device_id = ?", (device_id,)).fetchone()
    if not row:
        return jsonify({"error": "الجهاز غير موجود"}), 404
    return jsonify(device_to_dict(row, get_device_tags(db, device_id)))


@app.route("/api/devices/<device_id>", methods=["PUT"])
@login_required
def update_device(device_id):
    db = get_db()
    existing = db.execute("SELECT 1 FROM devices WHERE device_id = ?", (device_id,)).fetchone()
    if not existing:
        return jsonify({"error": "الجهاز غير موجود"}), 404

    data = request.get_json(silent=True) or {}

    def _to_float(v):
        try:
            return float(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    db.execute(
        "UPDATE devices SET description = ?, lat = ?, lng = ? WHERE device_id = ?",
        (
            data.get("description") or None,
            _to_float(data.get("lat")),
            _to_float(data.get("lng")),
            device_id,
        ),
    )

    # استبدال tags بالكامل بالقائمة الجديدة
    db.execute("DELETE FROM device_tags WHERE device_id = ?", (device_id,))
    tags = data.get("tags") or []
    for i, tag in enumerate(tags):
        name = (tag.get("tag_name") or "").strip()
        if not name:
            continue
        db.execute(
            "INSERT INTO device_tags (device_id, tag_name, unit, min_val, max_val, sort_order) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (
                device_id, name, (tag.get("unit") or "").strip() or None,
                _to_float(tag.get("min_val")), _to_float(tag.get("max_val")), i,
            ),
        )

    db.commit()
    return jsonify({"status": "ok"})


@app.route("/api/devices/<device_id>", methods=["DELETE"])
@login_required
def delete_device(device_id):
    db = get_db()
    db.execute("DELETE FROM devices WHERE device_id = ?", (device_id,))
    db.execute("DELETE FROM device_tags WHERE device_id = ?", (device_id,))
    db.execute("DELETE FROM readings WHERE device_id = ?", (device_id,))
    db.commit()
    return jsonify({"status": "ok"})


# ---------------------------------------------------------------------------
# API: استقبال القراءات (من ESP32)
# ---------------------------------------------------------------------------
@app.route("/api/reading", methods=["POST"])
def add_reading():
    """
    يستقبل قراءة جديدة من الجهاز. كل القيم تُرسل داخل "values"، بأسماء
    مطابقة لـ tags المعرّفة للجهاز في صفحة "إدارة الأجهزة".

    مثال (جهاز حرارة/رطوبة، مثل AM2302B):
    {
        "api_key": "changeme-esp32-key",
        "device_id": "am2302b-1",
        "values": {"temperature": 4.2, "humidity": 55.0}
    }

    مثال (جهاز VFD):
    {
        "api_key": "changeme-esp32-key",
        "device_id": "vfd-pump-1",
        "values": {"التردد": 45.2, "التيار": 3.1}
    }

    توافق قديم: يمكن أيضاً إرسال temperature/humidity كحقول مباشرة
    (بدون "values")، وسيتم دمجها تلقائياً.

    ملاحظة: الجهاز خاصو يكون مُضافاً مسبقاً من صفحة "إدارة الأجهزة"،
    وإلا السيرفر يرفض القراءة (404). التنبيه (alarm) يُحسب تلقائياً
    إذا أي قيمة خرجت عن Min/Max الخاص بـ tag ها من نفس الجهاز.
    """
    data = request.get_json(silent=True) or {}

    if data.get("api_key") != DEVICE_API_KEY:
        return jsonify({"error": "unauthorized"}), 401

    device_id = data.get("device_id")
    if not device_id:
        return jsonify({"error": "device_id مطلوب"}), 400

    db = get_db()
    device = db.execute("SELECT * FROM devices WHERE device_id = ?", (device_id,)).fetchone()
    if not device:
        return jsonify({"error": "الجهاز غير مسجّل. أضفه أولاً من صفحة إدارة الأجهزة"}), 404

    values = dict(data.get("values") or {})

    # توافق مع الصيغة القديمة (temperature/humidity كحقول مباشرة)
    if data.get("temperature") is not None:
        values.setdefault("temperature", data["temperature"])
    if data.get("humidity") is not None:
        values.setdefault("humidity", data["humidity"])

    if not values:
        return jsonify({"error": "يجب إرسال values على الأقل (أو temperature/humidity)"}), 400

    # تحويل كل القيم لأرقام حيثما أمكن
    clean_values = {}
    for k, v in values.items():
        try:
            clean_values[k] = float(v)
        except (TypeError, ValueError):
            clean_values[k] = v  # يُترك كما هو إذا لم يكن رقماً

    # حساب التنبيه: مرسل صراحة، أو تلقائياً إذا أي قيمة خرجت عن Min/Max الخاص بـ tag ها
    alarm = bool(data.get("alarm", False))
    tag_defs = {t["tag_name"]: t for t in get_device_tags(db, device_id)}
    for k, v in clean_values.items():
        if not isinstance(v, (int, float)):
            continue
        tag_def = tag_defs.get(k)
        if not tag_def:
            continue
        if tag_def["min_val"] is not None and v < tag_def["min_val"]:
            alarm = True
        if tag_def["max_val"] is not None and v > tag_def["max_val"]:
            alarm = True

    created_at = datetime.now(timezone.utc).isoformat()

    db.execute(
        """
        INSERT INTO readings (device_id, temperature, humidity, values_json, alarm, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            device_id,
            clean_values.get("temperature"),
            clean_values.get("humidity"),
            json.dumps(clean_values),
            1 if alarm else 0,
            created_at,
        ),
    )
    db.commit()

    return jsonify({"status": "ok", "alarm": alarm}), 201


# ---------------------------------------------------------------------------
# API: قراءة البيانات (مع فلترة زمنية)
# ---------------------------------------------------------------------------
def _apply_time_filters(query, params, device_id, date_from, date_to):
    query += " WHERE device_id = ?"
    params.append(device_id)
    if date_from:
        query += " AND created_at >= ?"
        params.append(date_from)
    if date_to:
        query += " AND created_at <= ?"
        params.append(date_to)
    return query, params


@app.route("/api/readings", methods=["GET"])
@login_required
def list_readings():
    device_id = request.args.get("device_id")
    if not device_id:
        return jsonify({"error": "device_id مطلوب"}), 400

    limit = request.args.get("limit", default=500, type=int)
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    db = get_db()
    query = "SELECT * FROM readings"
    params = []
    query, params = _apply_time_filters(query, params, device_id, date_from, date_to)
    query += " ORDER BY id DESC LIMIT ?"
    params.append(limit)

    rows = db.execute(query, params).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        try:
            d["values"] = json.loads(d["values_json"]) if d.get("values_json") else {}
        except (TypeError, json.JSONDecodeError):
            d["values"] = {}
        del d["values_json"]
        result.append(d)
    result.reverse()
    return jsonify(result)


@app.route("/api/readings/latest", methods=["GET"])
@login_required
def latest_reading():
    device_id = request.args.get("device_id")
    if not device_id:
        return jsonify({"error": "device_id مطلوب"}), 400

    db = get_db()
    row = db.execute(
        "SELECT * FROM readings WHERE device_id = ? ORDER BY id DESC LIMIT 1", (device_id,)
    ).fetchone()

    if not row:
        return jsonify(None)

    d = dict(row)
    try:
        d["values"] = json.loads(d["values_json"]) if d.get("values_json") else {}
    except (TypeError, json.JSONDecodeError):
        d["values"] = {}
    del d["values_json"]
    return jsonify(d)


# ---------------------------------------------------------------------------
# تصدير البيانات (CSV / Excel) — مع فلترة زمنية
# ---------------------------------------------------------------------------
EXPORT_TRANSLATIONS = {
    "ar": {"status": "الحالة", "datetime": "التاريخ والوقت", "temperature": "الحرارة °C",
           "humidity": "الرطوبة %", "alarm": "تنبيه", "normal": "طبيعي", "sheet_title": "القراءات"},
    "fr": {"status": "État", "datetime": "Date et heure", "temperature": "Temp. °C",
           "humidity": "Humidité %", "alarm": "Alarme", "normal": "Normal", "sheet_title": "Lectures"},
    "en": {"status": "Status", "datetime": "Date & Time", "temperature": "Temp °C",
           "humidity": "Humidity %", "alarm": "Alarm", "normal": "Normal", "sheet_title": "Readings"},
}


def get_export_lang():
    lang = request.args.get("lang", "ar")
    return lang if lang in EXPORT_TRANSLATIONS else "ar"


def build_export_table(rows, tr, tag_names):
    headers = tag_names + [tr["status"], tr["datetime"]]
    table_rows = []
    for r in rows:
        row = [r["values"].get(tag, "") for tag in tag_names]
        row.append(tr["alarm"] if r["alarm"] else tr["normal"])
        row.append(r["created_at"])
        table_rows.append(row)
    return headers, table_rows


def _fetch_export_rows(device_id, date_from, date_to):
    db = get_db()
    query = "SELECT * FROM readings"
    params = []
    query, params = _apply_time_filters(query, params, device_id, date_from, date_to)
    query += " ORDER BY id ASC"
    rows = db.execute(query, params).fetchall()

    result = []
    for r in rows:
        d = dict(r)
        try:
            d["values"] = json.loads(d["values_json"]) if d.get("values_json") else {}
        except (TypeError, json.JSONDecodeError):
            d["values"] = {}
        result.append(d)
    return result


@app.route("/api/export/csv", methods=["GET"])
@login_required
def export_csv():
    device_id = request.args.get("device_id")
    if not device_id:
        return jsonify({"error": "device_id مطلوب"}), 400

    lang = get_export_lang()
    tr = EXPORT_TRANSLATIONS[lang]
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    db = get_db()
    tag_names = [t["tag_name"] for t in get_device_tags(db, device_id)]
    rows = _fetch_export_rows(device_id, date_from, date_to)
    headers, table_rows = build_export_table(rows, tr, tag_names)

    output = io.StringIO()
    output.write("\ufeff")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(table_rows)

    filename = f"readings_{device_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/api/export/xlsx", methods=["GET"])
@login_required
def export_xlsx():
    device_id = request.args.get("device_id")
    if not device_id:
        return jsonify({"error": "device_id مطلوب"}), 400

    lang = get_export_lang()
    tr = EXPORT_TRANSLATIONS[lang]
    date_from = request.args.get("from")
    date_to = request.args.get("to")

    db = get_db()
    tag_names = [t["tag_name"] for t in get_device_tags(db, device_id)]
    rows = _fetch_export_rows(device_id, date_from, date_to)
    headers, table_rows = build_export_table(rows, tr, tag_names)

    wb = Workbook()
    ws = wb.active
    ws.title = tr["sheet_title"]
    ws.sheet_view.rightToLeft = (lang == "ar")

    header_fill = PatternFill(start_color="1F5A3B", end_color="1F5A3B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in table_rows:
        ws.append(row)

    for i, header in enumerate(headers, start=1):
        width = max(12, min(28, len(str(header)) + 6))
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"readings_{device_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
